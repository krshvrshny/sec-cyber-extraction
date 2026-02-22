import pandas as pd
import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────
PARQUET_PATH = "filings.parquet"
EXCEL_PATH   = "data_sample.xlsx"
OUTPUT_PATH  = "length_results"

# ── Load Metadata ─────────────────────────────────────────────────────────────
def load_metadata(excel_path):
    wb      = openpyxl.load_workbook(excel_path)
    ws      = wb.active
    headers = [cell.value for cell in ws[1]]

    metadata = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        ticker   = row_dict.get("ticker")
        if ticker:
            metadata[ticker] = {
                "size":       row_dict.get("size"),
                "market_cap": row_dict.get("market cap"),
            }
    return pd.DataFrame.from_dict(metadata, orient="index").reset_index().rename(columns={"index": "ticker"})


# ── Length Computation ────────────────────────────────────────────────────────
def compute_lengths(row):
    """
    Splits combined_text on the Item 1C delimiter to get separate
    word counts for 1A and 1C. Returns len_1a, len_1c, len_combined.
    """
    text   = row["combined_text"]
    has_1c = row["has_1c"]

    if not text:
        return 0, 0, 0

    if has_1c and "--- ITEM 1C ---" in text:
        parts  = text.split("--- ITEM 1C ---")
        len_1a = len(parts[0].split())
        len_1c = len(parts[1].split())
    else:
        len_1a = len(text.split())
        len_1c = 0

    return len_1a, len_1c, len_1a + len_1c


# ── Main ──────────────────────────────────────────────────────────────────────
def run_length_analysis(parquet_path=PARQUET_PATH, excel_path=EXCEL_PATH, output_path=OUTPUT_PATH):
    df       = pd.read_parquet(parquet_path)
    metadata = load_metadata(excel_path)

    # Merge in size and market cap
    df = df.merge(metadata, on="ticker", how="left")

    # Compute lengths
    df[["len_1a", "len_1c", "len_combined"]] = df.apply(
        compute_lengths, axis=1, result_type="expand"
    )

    # ── Parquet: keep combined_text for downstream pipeline steps ─────────────
    parquet_df = df[["ticker", "company_name", "sector", "year", "has_1c",
                     "size", "market_cap", "len_1a", "len_1c", "len_combined",
                     "combined_text"]]
    parquet_df.to_parquet(f"{output_path}.parquet", index=False)
    print(f"[INFO] Saved {output_path}.parquet")

    # ── CSV: only length-relevant columns, no combined_text ──────────────────
    csv_df = df[["ticker", "company_name", "sector", "year", "has_1c",
                 "size", "market_cap", "len_1a", "len_1c", "len_combined"]]
    csv_df.to_csv(f"{output_path}.csv", index=False)
    print(f"[INFO] Saved {output_path}.csv")

    # ── Summary CSVs ──────────────────────────────────────────────────────────
    summary_size = (
        df.groupby(["size", "year"])[["len_1a", "len_1c", "len_combined"]]
        .mean().round(0).reset_index()
    )
    summary_sector = (
        df.groupby(["sector", "year"])[["len_1a", "len_1c", "len_combined"]]
        .mean().round(0).reset_index()
    )

    summary_size.to_csv(f"{output_path}_by_size.csv", index=False)
    summary_sector.to_csv(f"{output_path}_by_sector.csv", index=False)
    print(f"[INFO] Saved {output_path}_by_size.csv")
    print(f"[INFO] Saved {output_path}_by_sector.csv")

    return df


if __name__ == "__main__":
    df = run_length_analysis()
    print("\n[INFO] Done.")
    print(df[["ticker", "sector", "year", "size", "len_1a", "len_1c", "len_combined"]].to_string())